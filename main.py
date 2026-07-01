import os
import asyncio
import logging
from telegram import (
    Update, ReplyKeyboardMarkup, ReplyKeyboardRemove,
    InlineKeyboardMarkup, InlineKeyboardButton
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ConversationHandler, filters, ContextTypes
)
from supabase import create_client

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

SB_URL = os.environ.get("SB_URL", "https://kaqfpzknyqcpdmalqhex.supabase.co")
SB_KEY = os.environ.get("SB_KEY")
TG_TOKEN = os.environ.get("TG_TOKEN")
ADMIN_CHAT_ID = int(os.environ.get("ADMIN_CHAT_ID", "-5218962193"))

db = create_client(SB_URL, SB_KEY)
WEB_URL = "https://cosportsreg.vercel.app"
PLATFORM_TIP = (
    "\n\n💡 Работай где удобно:\n"
    "• Telegram-бот — прямо здесь\n"
    f"• Мини-апп — t.me/cosports_reg_bot/csreg\n"
    f"• Веб-панель — {WEB_URL}"
)

# States
(AWAIT_COMPANY, AWAIT_PLAYER_NAME, AWAIT_PLAYER_PHONE,
 AWAIT_PLAYER_POSITION, AWAIT_PLAYER_EXTRA, CONFIRM_SUBMIT) = range(6)

# ─── HELPERS ─────────────────────────────────────────────────────────────────

async def get_hr(telegram_id: int):
    """Вернуть company_user + company если HR авторизован."""
    try:
        r = db.from_("company_users").select("*,companies(id,name)") \
            .eq("telegram_id", telegram_id).execute()
        return r.data[0] if r.data else None
    except Exception:
        return None

async def get_manager(telegram_id: int):
    """Вернуть admin_user если это менеджер или админ."""
    try:
        r = db.from_("admin_users").select("*")             .eq("telegram_id", telegram_id).execute()
        if r.data and r.data[0]["role"] in ("admin", "manager"):
            return r.data[0]
        return None
    except Exception:
        return None

async def get_registration(company_id: str, tournament_id: str):
    try:
        r = db.from_("team_registrations").select("*") \
            .eq("company_id", company_id).eq("tournament_id", tournament_id).execute()
        return r.data[0] if r.data else None
    except Exception:
        return None

async def ensure_registration(company_id: str, tournament_id: str):
    reg = await get_registration(company_id, tournament_id)
    if not reg:
        r = db.from_("team_registrations").insert({
            "company_id": company_id,
            "tournament_id": tournament_id,
            "status": "draft"
        }).execute()
        return r.data[0]
    return reg

async def open_tournaments():
    r = db.from_("tournaments").select("*").eq("status", "open").execute()
    return r.data or []

async def send_admin(app, text, reply_markup=None):
    try:
        await app.bot.send_message(ADMIN_CHAT_ID, text, reply_markup=reply_markup)
    except Exception as e:
        logger.warning(f"Admin notify failed: {e}")

def manager_menu():
    return ReplyKeyboardMarkup([
        ["🏆 Все турниры", "📋 Все заявки"],
        ["🔍 Найти команду", "👥 Игроки компании"],
        ["🌐 Открыть панель"]
    ], resize_keyboard=True)

def main_menu():
    return ReplyKeyboardMarkup([
        ["📋 Мои заявки", "👥 Состав команды"],
        ["➕ Добавить игрока", "📤 Отправить заявку"],
        ["📊 Статус заявки", "🔄 Сбросить сессию"],
        ["🌐 Открыть панель"]
    ], resize_keyboard=True)

# ─── /start ──────────────────────────────────────────────────────────────────

async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    # Проверяем QR-токен в аргументах /start
    args = ctx.args
    if args and len(args) == 1 and args[0].startswith("qr_"):
        token = args[0]
        tg_id = update.effective_user.id
        user = update.effective_user
        
        # Записываем telegram_id к токену
        try:
            result = db.from_("auth_tokens").select("*").eq("token", token).eq("used", False).execute()
            if result.data:
                db.from_("auth_tokens").update({
                    "telegram_id": tg_id,
                    "used": True
                }).eq("token", token).execute()
                
                await update.message.reply_text(
                    f"✅ Авторизация успешна!\n\n"
                    f"Вернись на сайт — он автоматически войдёт в систему.\n\n"
                    f"Или работай прямо здесь — напиши /start",
                    reply_markup=ReplyKeyboardRemove()
                )
                # Теперь обычный старт
                return await handle_start_logic(update, ctx)
            else:
                await update.message.reply_text("❌ Ссылка недействительна или устарела.")
        except Exception as e:
            logger.error(f"QR auth error: {e}")
            await update.message.reply_text("❌ Ошибка авторизации.")
        return

    return await handle_start_logic(update, ctx)

async def handle_start_logic(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    # 1. Проверяем менеджеров/админов
    manager = await get_manager(tg_id)
    if manager:
        ctx.user_data["role"] = manager["role"]
        ctx.user_data["manager_name"] = manager.get("full_name", "Менеджер")
        role_label = "👑 Администратор" if manager["role"] == "admin" else "🛠 Менеджер"
        await update.message.reply_text(
            f"👋 Привет, *{manager.get('full_name', '')}*! {role_label}\n\nВыбери действие:",
            parse_mode="Markdown",
            reply_markup=manager_menu()
        )
        return ConversationHandler.END

    # 2. Проверяем представителей команд
    hr = await get_hr(tg_id)

    if hr:
        ctx.user_data["company_id"] = hr["companies"]["id"]
        ctx.user_data["company_name"] = hr["companies"]["name"]
        ctx.user_data["role"] = "representative"
        await update.message.reply_text(
            f"👋 Привет! Ты представитель *{hr['companies']['name']}*.\n\n"
            f"Выбери действие или работай в веб-панели:\n{WEB_URL}",
            parse_mode="Markdown",
            reply_markup=main_menu()
        )
        return ConversationHandler.END

    # Проверить pending
    p = db.from_("pending_hr").select("*").eq("telegram_id", tg_id).execute()
    if p.data:
        p.data = p.data[0]
    if p.data:
        status = p.data["status"]
        if status == "pending":
            await update.message.reply_text(
                "⏳ Твоя заявка на доступ уже отправлена и ожидает подтверждения администратора."
            )
        elif status == "rejected":
            await update.message.reply_text(
                "❌ Твоя заявка была отклонена. Обратись к администратору КЛЧ."
            )
        return ConversationHandler.END

    # Новый пользователь
    await update.message.reply_text(
        "👋 Добро пожаловать в *CoReg КЛЧ*!\n\n"
        "Это система регистрации команд на турниры Корпоративной Лиги Чемпионов.\n\n"
        "Напиши название своей компании:",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )
    return AWAIT_COMPANY

async def receive_company(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    user = update.effective_user
    company_name = update.message.text.strip()

    db.from_("pending_hr").insert({
        "telegram_id": tg_id,
        "username": user.username or "",
        "full_name": user.full_name or "",
        "company_name": company_name,
        "status": "pending"
    }).execute()

    await update.message.reply_text(
        f"✅ Заявка отправлена!\n\n"
        f"Компания: *{company_name}*\n\n"
        f"Администратор рассмотрит заявку и даст тебе доступ. "
        f"Ты получишь уведомление здесь.",
        parse_mode="Markdown"
    )

    # Получить список компаний для кнопок
    companies_r = db.from_("companies").select("id,name").order("name").execute()
    companies_list = companies_r.data or []

    kb = []
    for c in companies_list[:10]:
        cname = c["name"][:20]
        kb.append([InlineKeyboardButton(
            f"✅ {cname}",
            callback_data=f"approve_hr:{tg_id}:{c['id']}"
        )])
    # Кнопка новой компании с именем что написал HR
    import urllib.parse
    safe_name = company_name[:30].replace(":", "-")
    kb.append([InlineKeyboardButton(
        f"➕ Новая: {safe_name[:20]}",
        callback_data=f"newco_hr:{tg_id}:{safe_name}"
    )])
    # Назначить менеджером
    kb.append([InlineKeyboardButton(
        "🛠 Назначить менеджером",
        callback_data=f"make_manager:{tg_id}"
    )])
    kb.append([InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_hr:{tg_id}")])

    await send_admin(
        ctx.application,
        f"🆕 Новая заявка на доступ!\n"
        f"👤 {user.full_name} (@{user.username or '—'})\n"
        f"🏢 Написал: {company_name}\n"
        f"🆔 Telegram ID: {tg_id}\n\n"
        f"Выбери действие:",
        reply_markup=InlineKeyboardMarkup(kb)
    )
    return ConversationHandler.END

# ─── МОИ ЗАЯВКИ ──────────────────────────────────────────────────────────────

async def my_registrations(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    hr = await get_hr(tg_id)
    if not hr:
        await update.message.reply_text("❌ Нет доступа. Напиши /start")
        return

    company_id = hr["companies"]["id"]
    r = db.from_("team_registrations").select("*,tournaments(name,sport_type)") \
        .eq("company_id", company_id).execute()
    regs = r.data or []

    if not regs:
        await update.message.reply_text(
            "У вас ещё нет заявок.\n\nОткрытые турниры:",
            reply_markup=main_menu()
        )
        tours = await open_tournaments()
        if tours:
            kb = [[InlineKeyboardButton(t["name"], callback_data=f"join_{t['id']}")] for t in tours]
            await update.message.reply_text(
                "Выбери турнир для регистрации:",
                reply_markup=InlineKeyboardMarkup(kb)
            )
        return

    sport_icon = {"chess":"♟️","esports_cs2":"🎮","bowling":"🎳","football":"⚽","table_tennis":"🏓"}
    status_text = {"draft":"📝 Черновик","submitted":"⏳ На проверке","approved":"✅ Принята","rejected":"❌ Отклонена"}

    text = f"📋 *Заявки команды {hr['companies']['name']}:*\n\n"
    for reg in regs:
        t = reg.get("tournaments", {})
        icon = sport_icon.get(t.get("sport_type",""), "🏆")
        st = status_text.get(reg["status"], reg["status"])
        text += f"{icon} {t.get('name','—')}\n{st}\n\n"

    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=main_menu())

async def join_tournament(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    tournament_id = query.data.replace("join_", "")

    tg_id = update.effective_user.id
    hr = await get_hr(tg_id)
    if not hr:
        await query.message.reply_text("❌ Нет доступа.")
        return

    company_id = hr["companies"]["id"]
    reg = await ensure_registration(company_id, tournament_id)
    ctx.user_data["company_id"] = company_id
    ctx.user_data["tournament_id"] = tournament_id
    ctx.user_data["reg_id"] = reg["id"]

    await query.message.reply_text(
        f"✅ Заявка на турнир создана!\n\nТеперь добавь игроков через «➕ Добавить игрока».",
        reply_markup=main_menu()
    )

# ─── СОСТАВ КОМАНДЫ ──────────────────────────────────────────────────────────

async def show_roster(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    hr = await get_hr(tg_id)
    if not hr:
        await update.message.reply_text("❌ Нет доступа.")
        return

    company_id = hr["companies"]["id"]

    # Выбрать активный турнир
    regs = db.from_("team_registrations").select("*,tournaments(name)") \
        .eq("company_id", company_id).eq("status", "draft").execute()
    if not regs.data:
        await update.message.reply_text("Нет активных заявок в статусе Черновик.")
        return

    reg = regs.data[0]
    roster = db.from_("tournament_roster").select("*,players_pool(full_name,phone)") \
        .eq("registration_id", reg["id"]).order("order_no").execute()

    if not roster.data:
        await update.message.reply_text(
            f"👥 Состав на *{reg['tournaments']['name']}*: пусто\n\nДобавь игроков через «➕ Добавить игрока»",
            parse_mode="Markdown"
        )
        return

    text = f"👥 *Состав на {reg['tournaments']['name']}:*\n\n"
    for i, r in enumerate(roster.data, 1):
        p = r.get("players_pool", {})
        role = "👑 Капитан" if r["order_no"] == 1 else ("🔄 Запасной" if r["is_reserve"] else "")
        text += f"{i}. {p.get('full_name','—')} {role}\n"

    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=main_menu())

# ─── ДОБАВИТЬ ИГРОКА ─────────────────────────────────────────────────────────

async def add_player_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    hr = await get_hr(tg_id)
    if not hr:
        await update.message.reply_text("❌ Нет доступа.")
        return ConversationHandler.END

    company_id = hr["companies"]["id"]
    ctx.user_data["company_id"] = company_id

    # Найти черновую заявку
    regs = db.from_("team_registrations").select("*,tournaments(name,roster_schema)") \
        .eq("company_id", company_id).eq("status", "draft").execute()
    if not regs.data:
        # Предложить турниры
        tours = await open_tournaments()
        if not tours:
            await update.message.reply_text("Нет открытых турниров.")
            return ConversationHandler.END
        kb = [[InlineKeyboardButton(t["name"], callback_data=f"join_{t['id']}")] for t in tours]
        await update.message.reply_text(
            "Сначала зарегистрируйся на турнир:",
            reply_markup=InlineKeyboardMarkup(kb)
        )
        return ConversationHandler.END

    reg = regs.data[0]
    ctx.user_data["reg_id"] = reg["id"]
    ctx.user_data["tournament_id"] = reg["tournament_id"]
    ctx.user_data["roster_schema"] = reg.get("tournaments", {}).get("roster_schema", {})
    ctx.user_data["new_player"] = {}

    # Проверить архив игроков
    pool = db.from_("players_pool").select("*").eq("company_id", company_id).eq("is_active", True).execute()
    pool_data = pool.data or []

    # Получить уже добавленных
    in_roster = db.from_("tournament_roster").select("player_id").eq("registration_id", reg["id"]).execute()
    in_ids = {r["player_id"] for r in (in_roster.data or [])}
    available = [p for p in pool_data if p["id"] not in in_ids]

    if available:
        kb = [[InlineKeyboardButton(p["full_name"], callback_data=f"pool_{p['id']}")] for p in available]
        kb.append([InlineKeyboardButton("➕ Новый игрок", callback_data="pool_new")])
        await update.message.reply_text(
            "Выбери игрока из архива или добавь нового:",
            reply_markup=InlineKeyboardMarkup(kb)
        )
    else:
        await update.message.reply_text(
            "Введи ФИО нового игрока:",
            reply_markup=ReplyKeyboardRemove()
        )
        ctx.user_data["adding_new"] = True
        return AWAIT_PLAYER_NAME

    return AWAIT_PLAYER_NAME

async def pool_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "pool_new":
        await query.message.reply_text("Введи ФИО нового игрока:")
        ctx.user_data["adding_new"] = True
        return AWAIT_PLAYER_NAME

    player_id = query.data.replace("pool_", "")
    pr = db.from_("players_pool").select("*").eq("id", player_id).execute()
    player = pr.data[0] if pr.data else {}

    # Добавить в состав
    roster = db.from_("tournament_roster").select("order_no") \
        .eq("registration_id", ctx.user_data["reg_id"]).order("order_no", desc=True).limit(1).execute()
    next_no = (roster.data[0]["order_no"] if roster.data else 0) + 1

    db.from_("tournament_roster").insert({
        "registration_id": ctx.user_data["reg_id"],
        "player_id": player_id,
        "order_no": next_no,
        "is_reserve": False
    }).execute()

    await query.message.reply_text(
        f"✅ *{player['full_name']}* добавлен в состав!",
        parse_mode="Markdown",
        reply_markup=main_menu()
    )
    return ConversationHandler.END

async def receive_player_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["new_player"]["name"] = update.message.text.strip()
    await update.message.reply_text("Телефон игрока (или напиши «-» если нет):")
    return AWAIT_PLAYER_PHONE

async def receive_player_phone(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    phone = update.message.text.strip()
    ctx.user_data["new_player"]["phone"] = None if phone == "-" else phone
    await update.message.reply_text("Должность в компании:")
    return AWAIT_PLAYER_POSITION

async def receive_player_position(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["new_player"]["position"] = update.message.text.strip()

    # Проверить дополнительные поля по schema
    schema = ctx.user_data.get("roster_schema", {})
    extra_fields = [f for f in schema.get("fields", []) if f["key"] not in ("position",)]
    ctx.user_data["extra_fields"] = extra_fields
    ctx.user_data["extra_idx"] = 0
    ctx.user_data["extra_data"] = {}

    if extra_fields:
        f = extra_fields[0]
        await update.message.reply_text(
            f"{f['label']}{'  *обязательно*' if f.get('required') else ' (необязательно, напиши «-»)'}:",
            parse_mode="Markdown"
        )
        return AWAIT_PLAYER_EXTRA

    return await save_new_player(update, ctx)

async def receive_player_extra(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    val = update.message.text.strip()
    fields = ctx.user_data["extra_fields"]
    idx = ctx.user_data["extra_idx"]
    f = fields[idx]

    if val != "-":
        ctx.user_data["extra_data"][f["key"]] = val

    idx += 1
    ctx.user_data["extra_idx"] = idx

    if idx < len(fields):
        nf = fields[idx]
        await update.message.reply_text(
            f"{nf['label']}{'  *обязательно*' if nf.get('required') else ' (необязательно, напиши «-»)'}:",
            parse_mode="Markdown"
        )
        return AWAIT_PLAYER_EXTRA

    return await save_new_player(update, ctx)

async def save_new_player(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    p = ctx.user_data["new_player"]
    company_id = ctx.user_data["company_id"]

    # Создать в players_pool
    new_p = db.from_("players_pool").insert({
        "company_id": company_id,
        "full_name": p["name"],
        "phone": p.get("phone"),
        "notes": p.get("position", "")
    }).execute()
    player_id = new_p.data[0]["id"]

    # Добавить в состав
    roster = db.from_("tournament_roster").select("order_no") \
        .eq("registration_id", ctx.user_data["reg_id"]).order("order_no", desc=True).limit(1).execute()
    next_no = (roster.data[0]["order_no"] if roster.data else 0) + 1

    db.from_("tournament_roster").insert({
        "registration_id": ctx.user_data["reg_id"],
        "player_id": player_id,
        "order_no": next_no,
        "is_reserve": False,
        "extra": ctx.user_data.get("extra_data", {})
    }).execute()

    await update.message.reply_text(
        f"✅ *{p['name']}* добавлен в состав и архив компании!"
        f"{PLATFORM_TIP}",
        parse_mode="Markdown",
        reply_markup=main_menu()
    )
    return ConversationHandler.END

# ─── ОТПРАВИТЬ ЗАЯВКУ ────────────────────────────────────────────────────────

async def submit_registration(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    hr = await get_hr(tg_id)
    if not hr:
        await update.message.reply_text("❌ Нет доступа.")
        return

    company_id = hr["companies"]["id"]
    regs = db.from_("team_registrations").select("*,tournaments(name,roster_min)") \
        .eq("company_id", company_id).eq("status", "draft").execute()
    if not regs.data:
        await update.message.reply_text("Нет заявок в статусе черновик.")
        return

    reg = regs.data[0]
    roster = db.from_("tournament_roster").select("id").eq("registration_id", reg["id"]).execute()
    count = len(roster.data or [])
    min_players = reg.get("tournaments", {}).get("roster_min", 1)

    if count < min_players:
        await update.message.reply_text(
            f"⚠️ Минимальный состав: {min_players} игроков. Сейчас добавлено: {count}."
        )
        return

    db.from_("team_registrations").update({
        "status": "submitted",
        "submitted_at": "now()"
    }).eq("id", reg["id"]).execute()

    await update.message.reply_text(
        f"📤 Заявка *{hr['companies']['name']}* на *{reg['tournaments']['name']}* отправлена на проверку!\n\n"
        f"Ты получишь уведомление когда администратор её рассмотрит."
        f"{PLATFORM_TIP}",
        parse_mode="Markdown",
        reply_markup=main_menu()
    )

    kb = [[
        InlineKeyboardButton("✅ Принять", callback_data=f"approve_reg:{reg['id']}"),
        InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_reg:{reg['id']}")
    ]]
    await send_admin(
        ctx.application,
        f"📋 Новая заявка на проверку!\n"
        f"🏢 {hr['companies']['name']}\n"
        f"🏆 {reg['tournaments']['name']}\n"
        f"👥 Игроков: {count}",
        reply_markup=InlineKeyboardMarkup(kb)
    )

# ─── СТАТУС ЗАЯВКИ ───────────────────────────────────────────────────────────

async def check_status(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    hr = await get_hr(tg_id)
    if not hr:
        await update.message.reply_text("❌ Нет доступа.")
        return

    company_id = hr["companies"]["id"]
    regs = db.from_("team_registrations").select("*,tournaments(name)") \
        .eq("company_id", company_id).order("created_at", desc=True).execute()

    if not regs.data:
        await update.message.reply_text("У вас ещё нет заявок.")
        return

    status_text = {
        "draft": "📝 Черновик — заявка не отправлена",
        "submitted": "⏳ На проверке у администратора",
        "approved": "✅ Заявка принята!",
        "rejected": "❌ Заявка отклонена"
    }
    text = f"📊 *Статус заявок {hr['companies']['name']}:*\n\n"
    for reg in regs.data:
        t = reg.get("tournaments", {})
        st = status_text.get(reg["status"], reg["status"])
        text += f"🏆 {t.get('name','—')}\n{st}\n"
        if reg.get("admin_comment"):
            text += f"💬 Комментарий: {reg['admin_comment']}\n"
        text += "\n"

    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=main_menu())

# ─── CANCEL ──────────────────────────────────────────────────────────────────

async def open_app(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Открыть мини-апп."""
    from telegram import WebAppInfo
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton(
            "🌐 Открыть панель CoReg",
            web_app=WebAppInfo(url="https://cosportsreg.vercel.app")
        )
    ]])
    await update.message.reply_text(
        "🌐 Открой панель регистрации прямо в Telegram:\n"
        "Или перейди по ссылке: t.me/cosports_reg_bot/csreg",
        reply_markup=kb
    )

async def cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Отменено.", reply_markup=main_menu())
    return ConversationHandler.END

# ─── ЗАГРУЗКА ФАЙЛА / ФОТО ───────────────────────────────────────────────────

async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Принять фото или файл, распознать текст через Claude Vision."""
    tg_id = update.effective_user.id
    hr = await get_hr(tg_id)
    if not hr:
        await update.message.reply_text("❌ Нет доступа.")
        return

    msg = await update.message.reply_text("🔍 Читаю документ...")

    try:
        import io, base64, os
        import httpx

        claude_key = os.environ.get("CLAUDE_KEY", "")
        if not claude_key:
            await msg.edit_text("⚠️ Claude API ключ не настроен.")
            return

        extracted_text = None
        is_image = False

        if update.message.photo:
            # Фото → Vision
            file = await update.message.photo[-1].get_file()
            file_bytes = await file.download_as_bytearray()
            b64 = base64.b64encode(file_bytes).decode()
            is_image = True
            media_type = "image/jpeg"

        elif update.message.document:
            file = await update.message.document.get_file()
            fname = (update.message.document.file_name or "").lower()
            file_bytes = await file.download_as_bytearray()

            if fname.endswith((".jpg", ".jpeg", ".png", ".webp")):
                # Изображение → Vision
                b64 = base64.b64encode(file_bytes).decode()
                is_image = True
                media_type = "image/jpeg" if not fname.endswith(".png") else "image/png"

            elif fname.endswith(".docx"):
                # DOCX → извлечь текст
                try:
                    from docx import Document
                    doc = Document(io.BytesIO(bytes(file_bytes)))
                    raw_text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
                    # Отправим текст в Claude для структурирования
                    payload_text = {
                        "model": "claude-sonnet-4-6",
                        "max_tokens": 1000,
                        "messages": [{"role": "user", "content": f"""Это заявка команды. Извлеки список игроков и верни ТОЛЬКО JSON без пояснений:
{{"captain": {{"name": "ФИО", "phone": "телефон", "position": "должность"}}, "vice_captain": {{"name": "ФИО", "phone": "телефон", "position": "должность"}}, "players": [{{"name": "ФИО", "phone": "телефон", "position": "должность"}}]}}
Если поле не найдено — пустая строка. Только JSON.

Текст документа:
{raw_text[:3000]}"""}]
                    }
                    async with httpx.AsyncClient(timeout=30) as client2:
                        r2 = await client2.post(
                            "https://api.anthropic.com/v1/messages",
                            headers={"x-api-key": claude_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                            json=payload_text
                        )
                        extracted_text = r2.json().get("content", [{}])[0].get("text", raw_text)
                except Exception as e:
                    extracted_text = f"Не удалось прочитать DOCX: {e}"

            elif fname.endswith(".pdf"):
                # PDF → конвертировать страницы в изображения и читать через Vision
                try:
                    import fitz  # pymupdf
                    doc_pdf = fitz.open(stream=bytes(file_bytes), filetype="pdf")
                    pages_text = []
                    for page_num in range(min(3, len(doc_pdf))):  # макс 3 страницы
                        page = doc_pdf[page_num]
                        # Сначала попробуем извлечь текст напрямую
                        page_text = page.get_text().strip()
                        if page_text:
                            pages_text.append(page_text)
                        else:
                            # Если текст не извлечь - рендерим как изображение
                            mat = fitz.Matrix(2, 2)
                            pix = page.get_pixmap(matrix=mat)
                            img_bytes = pix.tobytes("jpeg")
                            b64_page = base64.b64encode(img_bytes).decode()
                            payload_page = {
                                "model": "claude-sonnet-4-6",
                                "max_tokens": 800,
                                "messages": [{"role": "user", "content": [
                                    {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64_page}},
                                    {"type": "text", "text": "Прочитай весь текст на этой странице PDF. Отвечай на русском."}
                                ]}]
                            }
                            async with httpx.AsyncClient(timeout=30) as client:
                                r = await client.post(
                                    "https://api.anthropic.com/v1/messages",
                                    headers={"x-api-key": claude_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                                    json=payload_page
                                )
                                pages_text.append(r.json().get("content", [{}])[0].get("text", ""))
                    extracted_text = "\n\n--- Стр. ---\n\n".join(pages_text) if pages_text else "Текст не найден"
                except Exception as e:
                    extracted_text = f"Ошибка чтения PDF: {e}"

            elif fname.endswith((".xlsx", ".xls")):
                # Excel → извлечь данные
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(bytes(file_bytes)), data_only=True)
                    rows_text = []
                    for sheet in wb.sheetnames[:2]:  # макс 2 листа
                        ws = wb[sheet]
                        rows_text.append(f"📊 Лист: {sheet}")
                        for row in ws.iter_rows(max_row=50, values_only=True):
                            cells = [str(c) for c in row if c is not None]
                            if cells:
                                rows_text.append(" | ".join(cells))
                    extracted_text = "\n".join(rows_text) if rows_text else "Файл пуст"
                except Exception as e:
                    extracted_text = f"Ошибка чтения Excel: {e}"

            else:
                await msg.edit_text("❌ Поддерживаются: фото (jpg/png), Word (.docx), PDF, Excel (.xlsx).")
                return
        else:
            await msg.edit_text("❌ Неподдерживаемый тип файла.")
            return

        if is_image:
            # Отправить изображение в Claude Vision
            payload = {
                "model": "claude-sonnet-4-6",
                "max_tokens": 1000,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                        {"type": "text", "text": """Это заявка команды на соревнование. Извлеки список игроков и верни ТОЛЬКО JSON без пояснений:
{"captain": {"name": "ФИО", "phone": "телефон", "position": "должность"}, "vice_captain": {"name": "ФИО", "phone": "телефон", "position": "должность"}, "players": [{"name": "ФИО", "phone": "телефон", "position": "должность"}, ...]}
Если поле не найдено — пустая строка. Только JSON, без текста вокруг."""}
                    ]
                }]
            }
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": claude_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                    json=payload
                )
                result = resp.json()
            extracted_text = result.get("content", [{}])[0].get("text", "Не удалось прочитать")

        # Попробуем распарсить JSON с игроками
        import json, re
        parsed = None
        try:
            json_match = re.search(r'\{.*\}', extracted_text or '', re.DOTALL)
            if json_match:
                parsed = json.loads(json_match.group())
        except Exception:
            parsed = None

        if parsed:
            # Форматируем красиво
            lines = ["📋 Распознанные данные заявки:\n"]
            all_players = []

            cap = parsed.get("captain", {})
            if cap.get("name"):
                lines.append(f"👑 Капитан: {cap['name']}")
                if cap.get("position"): lines.append(f"   Должность: {cap['position']}")
                if cap.get("phone"): lines.append(f"   Телефон: {cap['phone']}")
                all_players.append({"name": cap["name"], "phone": cap.get("phone",""), "position": cap.get("position",""), "is_captain": True})

            vice = parsed.get("vice_captain", {})
            if vice.get("name"):
                lines.append(f"\n🥈 Вице-капитан: {vice['name']}")
                if vice.get("position"): lines.append(f"   Должность: {vice['position']}")
                if vice.get("phone"): lines.append(f"   Телефон: {vice['phone']}")
                all_players.append({"name": vice["name"], "phone": vice.get("phone",""), "position": vice.get("position",""), "is_captain": False})

            for i, p in enumerate(parsed.get("players", []), 1):
                if p.get("name") and p["name"] not in [x["name"] for x in all_players]:
                    lines.append(f"\n{i}. {p['name']}")
                    if p.get("position"): lines.append(f"   Должность: {p['position']}")
                    if p.get("phone"): lines.append(f"   Телефон: {p['phone']}")
                    all_players.append({"name": p["name"], "phone": p.get("phone",""), "position": p.get("position",""), "is_captain": False})

            lines.append(f"\n\nВсего игроков: {len(all_players)}")
            lines.append("\nВсё верно? Добавить всех в заявку?")

            # Сохраняем в ctx для подтверждения
            ctx.user_data["pending_players"] = all_players

            import json as json2
            kb = InlineKeyboardMarkup([[
                InlineKeyboardButton("✅ Да, добавить всех", callback_data="confirm_bulk_add"),
                InlineKeyboardButton("❌ Отмена", callback_data="cancel_bulk_add")
            ]])
            await msg.edit_text("\n".join(lines), reply_markup=kb)
        else:
            # Не удалось распарсить — показываем как текст
            safe_text = (extracted_text or "Пусто")[:2000]
            for ch in ['*', '_', '`', '[', ']', '#']:
                safe_text = safe_text.replace(ch, '')
            await msg.edit_text(f"📄 Результат:\n\n{safe_text}")

    except Exception as e:
        logger.error(f"Document processing error: {e}")
        await msg.edit_text(f"❌ Ошибка: {str(e)[:300]}")

# ─── MAIN ────────────────────────────────────────────────────────────────────

# ─── МЕНЕДЖЕР: ВСЕ ТУРНИРЫ ──────────────────────────────────────────────────

async def manager_all_tournaments(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    if not await get_manager(tg_id):
        await update.message.reply_text("❌ Нет доступа.")
        return

    tours = db.from_("tournaments").select("*").order("created_at", desc=True).execute()
    if not tours.data:
        await update.message.reply_text("Турниров нет.", reply_markup=manager_menu())
        return

    sport_icon = {"chess":"♟️","esports_cs2":"🎮","bowling":"🎳","football":"⚽","table_tennis":"🏓"}
    status_label = {"open":"🟢 Открыт","draft":"⚪️ Черновик","closed":"🔴 Закрыт"}
    text = "🏆 *Все турниры:*\n\n"
    kb = []
    for t in tours.data:
        icon = sport_icon.get(t["sport_type"], "🏆")
        st = status_label.get(t["status"], t["status"])
        text += f"{icon} {t['name']}\n{st}\n\n"
        kb.append([InlineKeyboardButton(
            f"{icon} {t['name'][:30]}",
            callback_data=f"mgr_tour:{t['id']}"
        )])

    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=manager_menu())
    await update.message.reply_text("Выбери турнир:", reply_markup=InlineKeyboardMarkup(kb))

async def manager_all_registrations(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    if not await get_manager(tg_id):
        await update.message.reply_text("❌ Нет доступа.")
        return

    regs = db.from_("team_registrations")         .select("*,companies(name),tournaments(name)")         .order("created_at", desc=True).execute()

    if not regs.data:
        await update.message.reply_text("Заявок нет.", reply_markup=manager_menu())
        return

    status_icon = {"draft":"📝","submitted":"⏳","approved":"✅","rejected":"❌"}
    text = "📋 *Все заявки:*\n\n"
    for r in regs.data[:20]:
        si = status_icon.get(r["status"], "❓")
        text += f"{si} {r['companies']['name']} → {r['tournaments']['name']}\n"

    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=manager_menu())

async def manager_tour_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    tournament_id = query.data.replace("mgr_tour:", "")

    regs = db.from_("team_registrations")         .select("*,companies(name),tournament_roster(id)")         .eq("tournament_id", tournament_id)         .order("created_at", desc=True).execute()

    tour = db.from_("tournaments").select("name").eq("id", tournament_id).execute()
    tour_name = tour.data[0]["name"] if tour.data else "—"

    if not regs.data:
        await query.message.reply_text(f"По турниру *{tour_name}* заявок нет.", parse_mode="Markdown")
        return

    status_icon = {"draft":"📝","submitted":"⏳","approved":"✅","rejected":"❌"}
    text = f"📋 *{tour_name}*\n\n"
    kb = []
    for r in regs.data:
        si = status_icon.get(r["status"], "❓")
        count = len(r.get("tournament_roster") or [])
        text += f"{si} {r['companies']['name']} — {count} игр.\n"
        if r["status"] == "submitted":
            kb.append([
                InlineKeyboardButton(f"✅ Принять {r['companies']['name'][:15]}", callback_data=f"approve_reg:{r['id']}"),
                InlineKeyboardButton("❌", callback_data=f"reject_reg:{r['id']}")
            ])

    await query.message.reply_text(text, parse_mode="Markdown")
    if kb:
        await query.message.reply_text("Заявки на рассмотрении:", reply_markup=InlineKeyboardMarkup(kb))

async def admin_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Обработка кнопок админа в групповом чате."""
    query = update.callback_query
    await query.answer()
    data = query.data

    # ── ПОДТВЕРДИТЬ HR ──────────────────────────────────────────────────────
    if data.startswith("approve_hr:"):
        _, tg_id_str, company_id = data.split(":", 2)
        tg_id = int(tg_id_str)

        # Обновить pending_hr
        db.from_("pending_hr").update({
            "status": "approved",
            "company_id": company_id
        }).eq("telegram_id", tg_id).execute()

        # Получить компанию
        comp = db.from_("companies").select("name").eq("id", company_id).execute()
        comp_name = comp.data[0]["name"] if comp.data else "—"

        # Найти или создать company_users запись с telegram_id
        existing = db.from_("company_users").select("id").eq("telegram_id", tg_id).execute()
        if not existing.data:
            # Создать новый логин на основе Telegram ID
            pending = db.from_("pending_hr").select("username,full_name").eq("telegram_id", tg_id).execute()
            p = pending.data[0] if pending.data else {}
            login = (p.get("username") or f"hr_{tg_id}").lower().replace(" ", "_")[:20]
            db.from_("company_users").insert({
                "company_id": company_id,
                "login": login,
                "password_hash": "tg_auth",
                "telegram_id": tg_id,
                "created_by": "admin_bot"
            }).execute()

        await query.edit_message_text(
            query.message.text + f"\n\n✅ Подтверждено → {comp_name}"
        )

        # Уведомить HR
        try:
            await ctx.bot.send_message(
                tg_id,
                f"✅ Твой доступ подтверждён!\n\n"
                f"🏢 Компания: *{comp_name}*\n\n"
                f"Напиши /start чтобы начать работу с заявкой.",
                parse_mode="Markdown"
            )
        except Exception as e:
            logger.warning(f"Cannot notify HR {tg_id}: {e}")

    # ── НОВАЯ КОМПАНИЯ + ПОДТВЕРДИТЬ HR ─────────────────────────────────────
    elif data.startswith("newco_hr:"):
        parts = data.split(":", 2)
        tg_id = int(parts[1])
        new_company_name = parts[2] if len(parts) > 2 else "Новая компания"

        # Создать компанию
        existing = db.from_("companies").select("id").eq("name", new_company_name).execute()
        if existing.data:
            company_id = existing.data[0]["id"]
        else:
            new_comp = db.from_("companies").insert({"name": new_company_name}).execute()
            company_id = new_comp.data[0]["id"]

        # Обновить pending_hr
        db.from_("pending_hr").update({
            "status": "approved",
            "company_id": company_id
        }).eq("telegram_id", tg_id).execute()

        # Создать company_user
        pending = db.from_("pending_hr").select("username,full_name").eq("telegram_id", tg_id).execute()
        p = pending.data[0] if pending.data else {}
        login = (p.get("username") or f"hr_{tg_id}").lower().replace(" ", "_")[:20]
        existing_cu = db.from_("company_users").select("id").eq("telegram_id", tg_id).execute()
        if not existing_cu.data:
            db.from_("company_users").insert({
                "company_id": company_id,
                "login": login,
                "password_hash": "tg_auth",
                "telegram_id": tg_id,
                "created_by": "admin_bot"
            }).execute()

        await query.edit_message_text(
            query.message.text + f"\n\n✅ Создана компания «{new_company_name}» и доступ выдан"
        )
        try:
            await ctx.bot.send_message(
                tg_id,
                f"✅ Доступ подтверждён!\n\n"
                f"🏢 Компания: *{new_company_name}*\n\n"
                f"Напиши /start чтобы начать работу.",
                parse_mode="Markdown"
            )
        except Exception as e:
            logger.warning(f"Cannot notify HR {tg_id}: {e}")

    # ── НАЗНАЧИТЬ МЕНЕДЖЕРОМ ─────────────────────────────────────────────────
    elif data.startswith("make_manager:"):
        _, tg_id_str = data.split(":", 1)
        tg_id = int(tg_id_str)

        pending = db.from_("pending_hr").select("*").eq("telegram_id", tg_id).execute()
        p = pending.data[0] if pending.data else {}
        full_name = p.get("full_name", f"Manager {tg_id}")
        username = p.get("username", "")
        login = (username or f"mgr_{tg_id}").lower()[:20]

        # Проверить не существует ли уже
        existing = db.from_("admin_users").select("id").eq("telegram_id", tg_id).execute()
        if not existing.data:
            db.from_("admin_users").insert({
                "full_name": full_name,
                "login": login,
                "password_hash": "tg_auth",
                "telegram_id": tg_id,
                "role": "manager"
            }).execute()

        db.from_("pending_hr").update({"status": "approved"}).eq("telegram_id", tg_id).execute()

        await query.edit_message_text(
            query.message.text + f"\n\n🛠 Назначен менеджером: {full_name}"
        )
        try:
            await ctx.bot.send_message(
                tg_id,
                f"✅ Тебе выдана роль *Менеджер КЛЧ*!\n\n"
                f"Напиши /start для входа в систему.",
                parse_mode="Markdown"
            )
        except Exception as e:
            logger.warning(f"Cannot notify manager {tg_id}: {e}")

    # ── ОТКЛОНИТЬ HR ────────────────────────────────────────────────────────
    elif data.startswith("reject_hr:"):
        _, tg_id_str = data.split(":", 1)
        tg_id = int(tg_id_str)

        db.from_("pending_hr").update({"status": "rejected"}).eq("telegram_id", tg_id).execute()

        await query.edit_message_text(query.message.text + "\n\n❌ Отклонено")

        try:
            await ctx.bot.send_message(
                tg_id,
                "❌ К сожалению, твоя заявка на доступ была отклонена.\n"
                "Обратись к организатору КЛЧ за дополнительной информацией."
            )
        except Exception as e:
            logger.warning(f"Cannot notify HR {tg_id}: {e}")

    # ── ПРИНЯТЬ ЗАЯВКУ КОМАНДЫ ───────────────────────────────────────────────
    elif data.startswith("approve_reg:"):
        _, reg_id = data.split(":", 1)
        db.from_("team_registrations").update({
            "status": "approved",
            "admin_comment": ""
        }).eq("id", reg_id).execute()

        # Найти HR компании
        reg = db.from_("team_registrations").select("*,companies(name),tournaments(name)") \
            .eq("id", reg_id).execute()
        if reg.data:
            r = reg.data[0]
            comp_id = r["company_id"]
            hr_user = db.from_("company_users").select("telegram_id").eq("company_id", comp_id).execute()
            if hr_user.data and hr_user.data[0].get("telegram_id"):
                try:
                    await ctx.bot.send_message(
                        hr_user.data[0]["telegram_id"],
                        f"✅ Ваша заявка *принята*!\n\n"
                        f"🏢 {r['companies']['name']}\n"
                        f"🏆 {r['tournaments']['name']}",
                        parse_mode="Markdown"
                    )
                except Exception as e:
                    logger.warning(f"Cannot notify HR: {e}")

        await query.edit_message_text(query.message.text + "\n\n✅ Заявка принята")

    # ── ОТКЛОНИТЬ ЗАЯВКУ КОМАНДЫ ─────────────────────────────────────────────
    elif data.startswith("reject_reg:"):
        _, reg_id = data.split(":", 1)
        db.from_("team_registrations").update({
            "status": "rejected",
            "admin_comment": "Отклонено администратором"
        }).eq("id", reg_id).execute()

        reg = db.from_("team_registrations").select("*,companies(name),tournaments(name)") \
            .eq("id", reg_id).execute()
        if reg.data:
            r = reg.data[0]
            comp_id = r["company_id"]
            hr_user = db.from_("company_users").select("telegram_id").eq("company_id", comp_id).execute()
            if hr_user.data and hr_user.data[0].get("telegram_id"):
                try:
                    await ctx.bot.send_message(
                        hr_user.data[0]["telegram_id"],
                        f"❌ Ваша заявка *отклонена*.\n\n"
                        f"🏢 {r['companies']['name']}\n"
                        f"🏆 {r['tournaments']['name']}\n\n"
                        f"Свяжитесь с организатором КЛЧ.",
                        parse_mode="Markdown"
                    )
                except Exception as e:
                    logger.warning(f"Cannot notify HR: {e}")

        await query.edit_message_text(query.message.text + "\n\n❌ Заявка отклонена")


async def bulk_add_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Подтверждение массового добавления игроков из документа."""
    query = update.callback_query
    await query.answer()

    if query.data == "cancel_bulk_add":
        await query.edit_message_text("❌ Отменено. Данные не добавлены.")
        return

    if query.data != "confirm_bulk_add":
        return

    tg_id = update.effective_user.id
    hr = await get_hr(tg_id)
    if not hr:
        await query.edit_message_text("❌ Нет доступа.")
        return

    players_to_add = ctx.user_data.get("pending_players", [])
    if not players_to_add:
        await query.edit_message_text("❌ Нет данных для добавления.")
        return

    company_id = hr["companies"]["id"]

    # Найти активную черновую заявку
    regs = db.from_("team_registrations").select("*,tournaments(roster_schema)")         .eq("company_id", company_id).eq("status", "draft").execute()
    if not regs.data:
        await query.edit_message_text("❌ Нет активной заявки в черновике. Сначала выбери турнир.")
        return

    reg = regs.data[0]
    reg_id = reg["id"]

    # Текущий максимальный order_no
    existing = db.from_("tournament_roster").select("order_no,player_id")         .eq("registration_id", reg_id).order("order_no", desc=True).execute()
    next_no = (existing.data[0]["order_no"] if existing.data else 0) + 1
    existing_player_ids = {r["player_id"] for r in (existing.data or [])}

    added = []
    skipped = []

    for p in players_to_add:
        if not p.get("name"):
            continue
        # Проверить есть ли уже в пуле
        pool_check = db.from_("players_pool").select("id")             .eq("company_id", company_id).eq("full_name", p["name"]).execute()

        if pool_check.data:
            player_id = pool_check.data[0]["id"]
            # Обновить телефон/должность если есть
            if p.get("phone") or p.get("position"):
                db.from_("players_pool").update({
                    "phone": p.get("phone") or None,
                    "notes": p.get("position") or None
                }).eq("id", player_id).execute()
        else:
            # Создать нового
            new_p = db.from_("players_pool").insert({
                "company_id": company_id,
                "full_name": p["name"],
                "phone": p.get("phone") or None,
                "notes": p.get("position") or None
            }).execute()
            player_id = new_p.data[0]["id"]

        # Добавить в состав если ещё нет
        if player_id not in existing_player_ids:
            is_captain = p.get("is_captain", False)
            order = 1 if is_captain else next_no
            if is_captain:
                # Сбросить текущего капитана
                db.from_("tournament_roster").update({"order_no": 99})                     .eq("registration_id", reg_id).eq("order_no", 1).execute()

            db.from_("tournament_roster").insert({
                "registration_id": reg_id,
                "player_id": player_id,
                "order_no": order,
                "is_reserve": False
            }).execute()
            if not is_captain:
                next_no += 1
            added.append(p["name"])
            existing_player_ids.add(player_id)
        else:
            skipped.append(p["name"])

    result = f"✅ Добавлено игроков: {len(added)}\n"
    if added:
        result += "\n".join(f"  • {n}" for n in added)
    if skipped:
        result += f"\n\n⏭ Уже были в заявке: {len(skipped)}"

    ctx.user_data.pop("pending_players", None)
    result += f"\n{PLATFORM_TIP}"
    await query.edit_message_text(result, reply_markup=None)

async def reset_session(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Сбросить сессию — удалить из pending_hr и начать заново."""
    tg_id = update.effective_user.id
    
    # Удалить из pending если есть
    db.from_("pending_hr").delete().eq("telegram_id", tg_id).execute()
    ctx.user_data.clear()
    
    await update.message.reply_text(
        "🔄 Сессия сброшена. Напиши /start чтобы начать заново.",
        reply_markup=ReplyKeyboardRemove()
    )

def main():
    app = Application.builder().token(TG_TOKEN).build()

    # Регистрация ConversationHandler
    reg_conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            AWAIT_COMPANY: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_company)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_user=True, per_chat=True
    )

    player_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^➕ Добавить игрока$"), add_player_start)],
        states={
            AWAIT_PLAYER_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_player_name),
                CallbackQueryHandler(pool_callback, pattern="^pool_"),
            ],
            AWAIT_PLAYER_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_player_phone)],
            AWAIT_PLAYER_POSITION: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_player_position)],
            AWAIT_PLAYER_EXTRA: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_player_extra)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_user=True, per_chat=True
    )

    app.add_handler(reg_conv)
    app.add_handler(player_conv)
    app.add_handler(CallbackQueryHandler(join_tournament, pattern="^join_"))
    app.add_handler(CallbackQueryHandler(pool_callback, pattern="^pool_"))
    app.add_handler(MessageHandler(filters.Regex("^📋 Мои заявки$"), my_registrations))
    app.add_handler(MessageHandler(filters.Regex("^👥 Состав команды$"), show_roster))
    app.add_handler(MessageHandler(filters.Regex("^📤 Отправить заявку$"), submit_registration))
    app.add_handler(MessageHandler(filters.Regex("^📊 Статус заявки$"), check_status))
    app.add_handler(MessageHandler(filters.Regex("^🔄 Сбросить сессию$"), reset_session))
    app.add_handler(CallbackQueryHandler(bulk_add_callback, pattern="^(confirm_bulk_add|cancel_bulk_add)$"))
    app.add_handler(MessageHandler(filters.PHOTO, handle_document))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(CallbackQueryHandler(admin_callback, pattern="^(approve_hr|reject_hr|approve_reg|reject_reg|newco_hr|make_manager):"))
    app.add_handler(CallbackQueryHandler(manager_tour_callback, pattern="^mgr_tour:"))
    app.add_handler(MessageHandler(filters.Regex("^🏆 Все турниры$"), manager_all_tournaments))
    app.add_handler(MessageHandler(filters.Regex("^📋 Все заявки$"), manager_all_registrations))

    app.add_handler(CommandHandler("reset", reset_session))
    app.add_handler(CommandHandler("app", open_app))
    app.add_handler(MessageHandler(filters.Regex("^🌐 Открыть панель$"), open_app))
    logger.info("CoReg bot started")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
